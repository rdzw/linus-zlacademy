# Importações de bibliotecas
from botcity.web import WebBot, Browser, By
from botcity.web.parsers import table_to_dict
from botcity.maestro import *
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date, timedelta
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
import matplotlib.pyplot as plt
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service

BotMaestroSDK.RAISE_NOT_CONNECTED = False

# Classe para gerenciar o Selenium
class SeleniumWrapper:
    def __init__(self):
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)

    def clear_field(self, field):
        field.clear()

    def close(self):
        self.driver.quit()

# Classe para gerenciar as operações com o BotCity
class BotCityWrapper(WebBot):
    def __init__(self):
        super().__init__()
        self.headless = False
        self.browser = Browser.CHROME
        self.driver_path = ChromeDriverManager().install()

    def clear_field(self, css_selector):
        selenium_wrapper = SeleniumWrapper()
        field = self.find_element(css_selector, By.CSS_SELECTOR)
        selenium_wrapper.clear_field(field)
        selenium_wrapper.close()
        
    def paste(self, css_selector, text):
        field = self.find_element(css_selector, By.CSS_SELECTOR)
        field.send_keys(text)

    def enter_iframe(self, iframe):
        self.driver.switch_to.frame(iframe)

# Classe principal para realizar o scraping e processamento de dados
class DollarScraperBot:
    def __init__(self):
        self.bot = BotCityWrapper()
        self.bot.headless = False
        self.bot.driver_path = ChromeDriverManager().install()
        
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.file_path = os.path.join(base_dir, "cotacao_dolar.xlsx")

    def run(self):
        # Configuração do BotCity
        maestro = BotMaestroSDK.from_sys_args()
        execution = maestro.get_execution()
        
        print(f"Task ID: {execution.task_id}")
        print(f"Task Parameters: {execution.parameters}")

        try:
            rates = self.get_dollar_rates()
            self.save_excel(self.file_path, rates)
            self.create_trend_graph(self.file_path)
            print("Processo concluído.")
            
            status = AutomationTaskFinishStatus.SUCCESS
            mensagem = "Tarefa BotMonitoraDolar finalizada com sucesso!"
        except Exception as e:
            
            self.bot.save_screenshot("erro.png")
            
            maestro.error(
                task_id=execution.task_id,
                exception=e,
                screenshot="erro.png"
            )
            
            status = AutomationTaskFinishStatus.FAILED
            mensagem = "Tarefa BotMonitoraDolar finalizada com erro!"
            
        finally:
            self.bot.stop_browser()
            
            maestro.finish_task(
                task_id=execution.task_id,
                status=status,
                mensagem=mensagem
            )

    def get_dollar_rates(self):
        url = "https://www.bcb.gov.br/estabilidadefinanceira/historicocotacoes"
        rates = []

        self.bot.browse(url)
        data_fim = date.today()
        data_inicio = data_fim - timedelta(days=30)

        data_fim_str = data_fim.strftime("%d/%m/%Y")
        data_inicio_str = data_inicio.strftime("%d/%m/%Y")
        
        rates = self.scrape_dollar_rate(data_inicio_str, data_fim_str)
        return rates

    def scrape_dollar_rate(self, data_inicio, data_fim):
        # Verificar se a página foi carregada
        while not self.bot.find_element('/html/body/app-root/bcb-cookies/div/div/div/div/button[2]', By.XPATH):
            print("A página ainda está carregando.")
            self.bot.wait(2000)
        
        self.bot.find_element('/html/body/app-root/bcb-cookies/div/div/div/div/button[2]', By.XPATH).click()
        iframe = self.bot.find_element('/html/body/app-root/app-root/div/div/main/dynamic-comp/div/div[2]/div[1]/div/iframe', By.XPATH)
        self.bot.enter_iframe(iframe)
        
        # Limpeza e preenchimento de campos usando SeleniumWrapper
        self.bot.clear_field("#DATAINI")
        print(f"Data início: {data_inicio}")
        self.bot.paste("#DATAINI", data_inicio.replace('/', ''))
        
        self.bot.clear_field("#DATAFIM")
        print(f"Data Fim: {data_fim}")
        self.bot.paste("#DATAFIM", data_fim.replace('/', ''))

        self.bot.find_element('/html/body/div/form/div/input', By.XPATH).click()
        
        # Verificar se a página foi carregada
        while not self.bot.find_element('/html/body/div[1]/table', By.XPATH):
            print("A página ainda está carregando.")
            self.bot.wait(2000)
            
        table_element = self.bot.find_element('/html/body/div[1]/table', By.XPATH)
        rate_data = table_to_dict(table_element, has_header=True)

        formatted_data = [{'date': row.get('data'), 'rate': float(row.get('cotações_em_real1').replace(',', '.'))}
                          for row in rate_data if 'data' in row and 'cotações_em_real1' in row]
        return formatted_data

    def save_excel(self, file_path, rates):
        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
        else:
            wb = Workbook()

        sheet = wb.active
        sheet.cell(row=1, column=1).value = "Data"
        sheet.cell(row=1, column=2).value = "Taxa do Dólar"
        for i, rate_data in enumerate(rates, start=2):
            sheet.cell(row=i, column=1).value = rate_data['date']
            sheet.cell(row=i, column=2).value = rate_data['rate']

        chart = BarChart()
        data = Reference(sheet, min_col=2, min_row=1, max_row=len(rates) + 1)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=len(rates) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = 'Variação Diária do Dólar'
        sheet.add_chart(chart, 'E2')

        color_scale = ColorScaleRule(start_type='min', start_color='FF99CC00', end_type='max', end_color='FFFF0000')
        sheet.conditional_formatting.add(f'B2:B{len(rates)+1}', color_scale)
        wb.save(file_path)

    def create_trend_graph(self, file_path):
        df = pd.read_excel(file_path)
        df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y')
        rates = df['Taxa do Dólar'].astype(float)

        plt.figure(figsize=(12, 6))
        plt.plot(df['Data'], rates, marker='o')
        plt.title('Tendência do Dólar ao Longo do Tempo')
        plt.xlabel('Data')
        plt.ylabel('Taxa do Dólar')
        plt.grid(True)
        plt.tight_layout()

        graph_file = os.path.splitext(file_path)[0] + '_trend.png'
        plt.savefig(graph_file)
        plt.close()

# Execução principal
if __name__ == "__main__":
    scraper_bot = DollarScraperBot()
    scraper_bot.run()

